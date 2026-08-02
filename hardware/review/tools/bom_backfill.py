"""Backfill real LCSC part numbers onto every BOM line that said 'generic'.

Part numbers researched against live LCSC/JLC stock data 2026-08-02. Selection rules
were: prefer JLC basic/preferred library; 0603 1% for resistors; X7R/X5R/C0G only for
MLCC (no Y5V/Z5U); >=25 V on the 9-18 V analog rail; electrolytic case size matched
EXACTLY to the footprint; stock >= 3000.

Keyed on (normalised value, footprint) because "22u" is an 0805 MLCC in one place and a
5 x 5.4 mm electrolytic in another, and "10u" likewise.

Usage: bom_backfill.py [--apply]      (dry run by default; writes a .bak first)
"""
import re
import shutil
import sys

import openpyxl

SRC = r"C:\Users\Jason\source\repos\Glitchwave\hardware\BOM.xlsx"
BAK = SRC.replace(".xlsx", ".bak_backfill_20260802.xlsx")

R = "R_0603_1608Metric"
C6 = "C_0603_1608Metric"

# (value, footprint-tail) -> (MPN, LCSC)
MAP = {
    # ---- 0603 1% resistors, UNI-ROYAL 0603WAF series except where noted ----
    ("0R", R): ("0603WAF0000T5E", "C21189"),
    ("100R", R): ("0603WAF1000T5E", "C22775"),
    ("470R", R): ("0603WAF4700T5E", "C23179"),
    ("680", R): ("0603WAF6800T5E", "C23228"),
    ("845R", R): ("FRC0603F8450TS", "C2933262"),      # FOJAN - no UNI-ROYAL 845R exists
    ("1k", R): ("0603WAF1001T5E", "C21190"),
    ("2.2k", R): ("0603WAF2201T5E", "C4190"),
    ("3.3k", R): ("0603WAF3301T5E", "C22978"),
    ("3.6k", R): ("0603WAF3601T5E", "C22980"),
    ("3.9k", R): ("0603WAF3901T5E", "C23018"),
    ("4.3k", R): ("0603WAF4301T5E", "C23159"),
    ("4.7k", R): ("0603WAF4701T5E", "C23162"),
    ("6.8k", R): ("0603WAF6801T5E", "C23212"),
    ("7k68", R): ("0603WAF7681T5E", "C23102"),
    ("10k", R): ("0603WAF1002T5E", "C25804"),
    ("15k", R): ("0603WAF1502T5E", "C22809"),
    ("18k", R): ("0603WAF1802T5E", "C25810"),
    ("20k", R): ("0603WAF2002T5E", "C4184"),
    ("22k", R): ("0603WAF2202T5E", "C31850"),
    ("24k9", R): ("0603WAF2492T5E", "C25962"),
    ("33k", R): ("0603WAF3302T5E", "C4216"),
    ("36k", R): ("0603WAF3602T5E", "C23147"),
    ("39k", R): ("0603WAF3902T5E", "C23153"),
    ("40k2", R): ("0603WAF4022T5E", "C12447"),
    ("43k", R): ("0603WAF4302T5E", "C23172"),
    ("47k", R): ("0603WAF4702T5E", "C25819"),
    ("51k", R): ("0603WAF5102T5E", "C23196"),
    ("100k", R): ("0603WAF1003T5E", "C25803"),
    ("147k", R): ("0603WAF1473T5E", "C22878"),
    ("150k", R): ("0603WAF1503T5E", "C22807"),
    ("200k", R): ("0603WAF2003T5E", "C25811"),
    ("220k", R): ("0603WAF2203T5E", "C22961"),
    ("330k", R): ("0603WAF3303T5E", "C23137"),
    ("470k", R): ("0603WAF4703T5E", "C23178"),
    ("1M", R): ("0603WAF1004T5E", "C22935"),
    ("2.2M", R): ("0603WAF2204T5E", "C22938"),
    # ---- MLCC 0603 ----
    ("220p", C6): ("CL10C221JB8NNNC", "C27675"),       # C0G 50V - filter position
    ("1n", C6): ("CL10C102JB8NNNC", "C163508"),        # C0G 50V
    ("3n3", C6): ("0603N332J500CT", "C152910"),        # NP0 50V
    ("3.9n", C6): ("GRM1885C1H392JA01D", "C415537"),   # C0G 50V
    ("10n", C6): ("GRM1885C1H103JA01D", "C85973"),     # C0G 50V
    ("47n", C6): ("CL10B473KB8NNNC", "C1622"),         # X7R 50V basic
    ("56n", C6): ("CC0603KRX7R9BB563", "C282072"),     # X7R 50V
    ("100n", C6): ("CC0603KRX7R9BB104", "C14663"),     # X7R 50V basic
    ("220n", C6): ("TCC0603X7R224K500CT", "C344195"),  # X7R 50V - see note
    ("1u", C6): ("CL10B105KB8NQNC", "C5199872"),       # X7R 50V - see note
    # ---- MLCC larger ----
    ("22u", "C_0805_2012Metric"): ("CL21A226MAQNNNE", "C45783"),      # X5R 25V basic
    ("10u/50V X7R", "C_1206_3216Metric"): ("CL31B106KBHNNNE", "C89632"),  # X7R 50V
    # ---- aluminium electrolytic, case size matched exactly ----
    ("470u", "CP_Elec_10x10.2"): ("RVT1V471M1010", "C3350"),
    ("220u/35V", "CP_Elec_8x10.2"): ("RVT1V221M0810", "C3340"),
    ("100u", "CP_Elec_6.3x7.7"): ("RVE100UF35V67RV0072", "C2836437"),
    ("47u", "CP_Elec_6.3x5.4"): ("RVT1V470M0605", "C72522"),
    ("22u", "CP_Elec_5x5.4"): ("RVT1V220M0505", "C72504"),
    ("10u", "CP_Elec_4x5.4"): ("RVT1V100M0405", "C72485"),
    ("4.7u", "CP_Elec_4x5.4"): ("RVT1V4R7M0405", "C72511"),
    # ---- ferrite ----
    ("600R/3A", "L_1206_3216Metric"): ("HCB3216KF-601T30", "C357023"),
    # ---- board-to-board stack pair, 2.5 mm insulator + 8.5 mm body = 11.00 mm gap ----
    # MAIN gets the SOCKET: it carries the DC jack, so the live half stays recessed.
    ("CTRL 2x8", "PinHeader_2x08_P2.54mm_Vertical"): ("PM254V-12-16-H85", "C46595985"),
    # CONTROL gets the HEADER, bottom side: lighter half on the hanging board.
    ("to MAIN 2x8", "PinHeader_2x08_P2.54mm_Vertical"): ("PZ254V-12-16P", "C492425"),
}

# DNP lines: no part to buy.
DNP_VALUES = {"DNP-LFIL", "DNP-OFIL"}


def norm(v):
    """'24k9 (EN divider - abs max 6V!)' -> '24k9';  '40k2 1%' -> '40k2'."""
    s = str(v or "").strip()
    s = re.sub(r"\(.*?\)", "", s).strip()
    s = re.sub(r"\s+1%$", "", s).strip()
    return s


def main():
    apply_it = "--apply" in sys.argv
    wb = openpyxl.load_workbook(SRC)
    changed, missed = [], []

    for sheet in ("MAIN board", "CONTROL board"):
        ws = wb[sheet]
        hdr = [str(c.value or "").strip() for c in ws[1]]
        i_val = hdr.index("Value")
        i_fp = hdr.index("Footprint")
        i_mpn = hdr.index("MPN")
        i_pn = hdr.index("Part # (LCSC C# / DK)")
        i_ref = hdr.index("Refs")
        for row in ws.iter_rows(min_row=2):
            if not row[i_ref].value:
                continue
            pn = str(row[i_pn].value or "").strip()
            if pn.lower() != "generic":
                continue
            val = norm(row[i_val].value)
            fp = str(row[i_fp].value or "").split(":")[-1]
            if val in DNP_VALUES:
                changed.append((sheet, row[i_ref].value, val, "DNP - not fitted", "DNP"))
                if apply_it:
                    row[i_mpn].value = "DNP - not fitted"
                    row[i_pn].value = "DNP"
                continue
            hit = MAP.get((val, fp)) or MAP.get((str(row[i_val].value).strip(), fp))
            if not hit:
                missed.append((sheet, row[i_ref].value, val, fp))
                continue
            mpn, lcsc = hit
            changed.append((sheet, row[i_ref].value, val, mpn, lcsc))
            if apply_it:
                row[i_mpn].value = mpn
                row[i_pn].value = lcsc

    print("resolved %d lines" % len(changed))
    for s, ref, val, mpn, lcsc in changed:
        print("  %-5s %-22s %-14s -> %-22s %s" % (
            s[:5], str(ref)[:22], val, mpn, lcsc))
    if missed:
        print()
        print("STILL UNRESOLVED (%d):" % len(missed))
        for s, ref, val, fp in missed:
            print("  %-5s %-22s value=%-16r footprint=%r" % (s[:5], str(ref)[:22], val, fp))

    if not apply_it:
        print("\nDRY RUN - nothing written.")
        return 1 if missed else 0
    if missed:
        print("\nABORT: unresolved lines above. Nothing written.")
        return 1
    shutil.copy2(SRC, BAK)
    wb.save(SRC)
    print("\nbackup: %s\nWROTE:  %s" % (BAK, SRC))
    return 0


if __name__ == "__main__":
    sys.exit(main())
