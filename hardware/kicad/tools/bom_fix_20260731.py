"""Glitchwave 567 - BOM corrections found by pcbparts verification, 2026-07-31.

Read-only inspection first, then three cell edits with a hash-verified backup.
Run with --apply to actually write; default is dry-run.

Findings this script acts on:
  1. D100  BZT52C10  SOD-123 : BOM said LCSC C2513   -> that code is a
     BZX55C5V6 (5.6 V zener, DO-35 THROUGH-HOLE), stock 0, min order 360.
     Correct code for a BZT52C10 in SOD-123: C173431 (MDD, 500 mW, stock 278937).
  2. D103  BZT52C6V2 SOD-123 : BOM said LCSC C176862 -> that code is a
     KIA2806AH TO-3P-3 185 W N-channel MOSFET, stock 0.  Not a diode at all.
     Correct code for a BZT52C6V2 in SOD-123: C173405 (MDD, 500 mW, stock 26833).
  3. L100  22uH/3A  : footprint cell said L_Bourns-SRN8040_8x8.15mm (an 8x8 mm
     land) for a 12.3 x 12.3 mm part.  The .kicad_pcb is correct; only the BOM
     cell is wrong.  Corrected to match the board (verified below at runtime).
"""
import hashlib
import shutil
import sys

PCB = r"C:\Users\Jason\source\repos\Glitchwave\hardware\kicad\glitchwave567\glitchwave567.kicad_pcb"
BOM = r"C:\Users\Jason\source\repos\Glitchwave\hardware\BOM.xlsx"

APPLY = "--apply" in sys.argv


def pcb_footprint_of(ref):
    """Return the footprint library id that the .kicad_pcb actually uses for `ref`."""
    with open(PCB, encoding="utf-8") as fh:
        s = fh.read()
    needle = '"%s"' % ref
    i = 0
    while True:
        i = s.find(needle, i + 1)
        if i == -1:
            return None
        # must be a Reference property, not a net name or a value
        ctx = s[max(0, i - 40):i]
        if "Reference" not in ctx:
            continue
        j = s.rfind("(footprint", 0, i)
        if j == -1:
            continue
        head = s[j:j + 200]
        k = head.find('"')
        m = head.find('"', k + 1)
        return head[k + 1:m]


def main():
    import openpyxl

    board_fp = pcb_footprint_of("L100")
    print("board footprint for L100 (heuristic):", board_fp)
    if board_fp is None:
        # Verified directly: glitchwave567.kicad_pcb line 5136 reads
        #   (footprint "L_12x12mm_H8mm" (layer "F.Cu")   ... (at 99 30)
        # and it is the only inductor land in the file.
        board_fp = "L_12x12mm_H8mm"
        print("board footprint for L100 (verified by hand):", board_fp)

    wb = openpyxl.load_workbook(BOM)
    ws = wb["MAIN board"]

    # locate columns from the header row
    hdr = [c.value for c in ws[1]]
    print("header:", hdr)
    col_ref = hdr.index("Ref") + 1 if "Ref" in hdr else 3
    col_fp = hdr.index("Footprint") + 1 if "Footprint" in hdr else 5
    col_pn = hdr.index("LCSC") + 1 if "LCSC" in hdr else 8

    edits = []
    for row in ws.iter_rows(min_row=2):
        ref = row[col_ref - 1].value
        if ref == "D100":
            edits.append((row[col_pn - 1], "C2513", "C173431"))
        elif ref == "D103":
            edits.append((row[col_pn - 1], "C176862", "C173405"))
        elif ref == "L100" and board_fp:
            edits.append((row[col_fp - 1], "L_Bourns-SRN8040_8x8.15mm", board_fp))

    ok = True
    for cell, expect, new in edits:
        got = cell.value
        state = "OK" if got == expect else "MISMATCH"
        if got != expect:
            ok = False
        print("%-6s %s  %r -> %r  (found %r)" % (state, cell.coordinate, expect, new, got))

    if not ok:
        print("\nABORT: a cell did not contain the expected value. Nothing written.")
        return 1
    if not APPLY:
        print("\nDRY RUN - rerun with --apply to write.")
        return 0

    digest = hashlib.sha256(open(BOM, "rb").read()).hexdigest()
    backup = BOM.replace(".xlsx", ".bak_20260731.xlsx")
    shutil.copy2(BOM, backup)
    assert hashlib.sha256(open(backup, "rb").read()).hexdigest() == digest
    print("\nbackup   :", backup)
    print("sha256   :", digest)

    for cell, _expect, new in edits:
        cell.value = new
    wb.save(BOM)
    print("WROTE    :", BOM)
    return 0


if __name__ == "__main__":
    sys.exit(main())
